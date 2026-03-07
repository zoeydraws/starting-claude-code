"""
Anonymize Research Data
=======================
Removes personally identifiable information (PII) from research files
using Microsoft Presidio.

Supported file types: .md, .txt, .csv, .xlsx, .json

Usage:
    1. Drop files into the input/ folder
    2. Run: python3 anonymize.py
       Or for specific files: python3 anonymize.py filename.md
    3. Find anonymized copies in the output/ folder

Replaces PII with labels like <LOCATION>, <EMAIL_ADDRESS>, etc.
People get numbered labels so you can track who said what:
  e.g. Jane Doe → <PERSON_1>, John Smith → <PERSON_2>, etc.

Original files are never modified.
A person mapping file (entity_map.txt) is saved to the output folder for reference.

Setup (one-time):
    pip3 install 'presidio-analyzer[transformers]' openpyxl protobuf sentencepiece 'transformers<4.57'
    python3 -m spacy download en_core_web_sm
"""

import re
import sys
import csv
import json
import shutil
from pathlib import Path

try:
    from presidio_analyzer import AnalyzerEngine
    from presidio_analyzer.nlp_engine import TransformersNlpEngine, NerModelConfiguration
except ImportError:
    print("Presidio is not installed. Run this command first:\n")
    print("  pip3 install 'presidio-analyzer[transformers]' openpyxl protobuf sentencepiece 'transformers<4.57' && python3 -m spacy download en_core_web_sm")
    sys.exit(1)

# --- Configuration ---

# Folder paths (relative to this script)
SCRIPT_DIR = Path(__file__).parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"

# File types to process
SUPPORTED_EXTENSIONS = {".md", ".txt", ".csv", ".xlsx", ".json"}

# PII entity types to detect (Presidio built-in)
# Full list: https://microsoft.github.io/presidio/supported_entities/
ENTITIES_TO_DETECT = [
    "PERSON",
    "EMAIL_ADDRESS",
    "PHONE_NUMBER",
    "IP_ADDRESS",
    "CREDIT_CARD",
    "SG_NRIC_FIN",
]

# Minimum confidence score (0.0 to 1.0) – lower = more aggressive detection
# 0.4 catches more but may have false positives; 0.7 is more conservative
SCORE_THRESHOLD = 0.4


# --- Engine setup ---

print("Loading Presidio engines (multilingual transformer model)...")

# Use a multilingual transformer model for NER instead of spaCy's en_core_web_lg.
# This recognizes non-Western names (Malay, Chinese, Indian, etc.) much better.
# spaCy en_core_web_sm is used only for tokenization, not NER.
ner_config = NerModelConfiguration(
    model_to_presidio_entity_mapping={
        "PER": "PERSON",
        "LOC": "LOCATION",
        "ORG": "ORGANIZATION",
    },
    labels_to_ignore=["O"],
    aggregation_strategy="simple",
    alignment_mode="expand",
)

nlp_engine = TransformersNlpEngine(
    models=[{
        "lang_code": "en",
        "model_name": {
            "spacy": "en_core_web_sm",
            "transformers": "Davlan/xlm-roberta-large-ner-hrl",
        },
    }],
    ner_model_configuration=ner_config,
)

analyzer = AnalyzerEngine(nlp_engine=nlp_engine)
print("Ready.\n")

# Global mapping: tracks assigned numbers per entity type across all files
# Key: (entity_type, original_text) → Value: numbered label like "PERSON_1"
entity_map = {}
# Counter per entity type: {"PERSON": 2, "LOCATION": 3, ...}
entity_counters = {}
# Max confidence score seen for each entity: ("PERSON", "Aniha") → 0.95
# Used by cleanup pass to only propagate high-confidence names
entity_scores = {}

# Minimum score for a name to be propagated in the cleanup pass.
# Names below this threshold (e.g. "Great", "So") won't be spread to other lines.
CLEANUP_SCORE_THRESHOLD = 0.7


def get_label(entity_type: str, original_text: str, score: float = 1.0) -> str:
    """Get or create a label for a detected entity.

    PERSON entities get numbered labels (PERSON_1, PERSON_2) so you can track
    who said what. All other entity types get a generic label (LOCATION, EMAIL_ADDRESS).

    For PERSON, first names are linked to their full name if one was seen earlier
    (e.g. "Michelle" reuses the label from "Michelle Park").
    """
    # Non-person entities: just return the type name, no numbering
    if entity_type != "PERSON":
        return entity_type

    text = original_text.strip()
    key = (entity_type, text)

    # Track the highest confidence score seen for this name
    entity_scores[key] = max(entity_scores.get(key, 0), score)

    if key not in entity_map:
        # Check if this is a first name of an already-seen full name
        # (or a full name whose first name was already seen)
        parts = text.split()
        if len(parts) == 1:
            for (etype, existing_text), label in entity_map.items():
                if etype == "PERSON" and existing_text.split()[0] == text:
                    entity_map[key] = label
                    return label
        else:
            first_name_key = ("PERSON", parts[0])
            if first_name_key in entity_map:
                entity_map[key] = entity_map[first_name_key]
                return entity_map[first_name_key]

        count = entity_counters.get(entity_type, 0) + 1
        entity_counters[entity_type] = count
        entity_map[key] = f"{entity_type}_{count}"

    return entity_map[key]


def find_speaker_labels(text: str) -> list[dict]:
    """Find names in markdown speaker labels like **Name:** that Presidio misses.

    Returns a list of detections with start/end positions and the name text,
    in the same shape as Presidio results so they can be merged.
    """
    speaker_pattern = re.compile(r"(?m)^\*\*([A-Z][a-z]+(?:\s[A-Z][a-z]+)*):\*\*")
    detections = []

    for match in speaker_pattern.finditer(text):
        name = match.group(1)
        # Verify with Presidio that it's actually a person name
        results = analyzer.analyze(
            text=f"My name is {name}.",
            entities=["PERSON"],
            language="en",
            score_threshold=SCORE_THRESHOLD,
        )
        if results:
            # Position of the name itself (inside the ** **)
            name_start = match.start() + 2  # skip leading **
            name_end = name_start + len(name)
            detections.append({
                "entity_type": "PERSON",
                "start": name_start,
                "end": name_end,
                "text": name,
            })

    return detections


def anonymize_line(line: str) -> str:
    """Detect all PII in a single line and replace with labeled placeholders.

    Combines Presidio detection with custom speaker-label detection, then
    replaces each unique entity with a tracked label (e.g. PERSON_1).
    """
    if not line or not line.strip():
        return line

    # Collect detections from Presidio
    presidio_results = analyzer.analyze(
        text=line,
        entities=ENTITIES_TO_DETECT,
        language="en",
        score_threshold=SCORE_THRESHOLD,
    )

    detections = []
    for r in presidio_results:
        detections.append({
            "entity_type": r.entity_type,
            "start": r.start,
            "end": r.end,
            "text": line[r.start:r.end],
            "score": r.score,
        })

    # Add speaker label detections that Presidio missed
    speaker_detections = find_speaker_labels(line)
    for sd in speaker_detections:
        overlaps = any(
            d["start"] <= sd["start"] and d["end"] >= sd["end"]
            for d in detections
        )
        if not overlaps:
            detections.append(sd)

    if not detections:
        return line

    # Deduplicate overlapping spans: keep the longer detection when two overlap
    # (e.g. an email "x@y.com" might also match as URL "y.com" – keep the email)
    detections.sort(key=lambda d: (d["start"], -(d["end"] - d["start"])))
    deduped = []
    for d in detections:
        if deduped and d["start"] < deduped[-1]["end"]:
            # Overlaps with previous – skip (previous is longer or same start)
            continue
        deduped.append(d)

    # Sort by start position descending so we can replace from end to start
    # (this preserves earlier character positions as we modify the string)
    deduped.sort(key=lambda d: d["start"], reverse=True)

    # Replace each detection with its label
    for d in deduped:
        label = get_label(d["entity_type"], d["text"], d.get("score", 1.0))
        line = line[:d["start"]] + f"<{label}>" + line[d["end"]:]

    return line


def cleanup_pass(text: str) -> str:
    """Second pass: catch any remaining occurrences of names the NER missed.

    Only propagates names that the transformer was confident about (score above
    CLEANUP_SCORE_THRESHOLD), so false positives like "Great" or "So" don't spread.
    """
    # Collect high-confidence PERSON names from entity_map
    names_to_labels = {}
    for (etype, name), label in entity_map.items():
        if etype != "PERSON":
            continue
        score = entity_scores.get((etype, name), 0)
        if score < CLEANUP_SCORE_THRESHOLD:
            continue
        names_to_labels[name] = label

    if not names_to_labels:
        return text

    # Sort by length descending so "Kwee Seng" is matched before "Kwee"
    sorted_names = sorted(names_to_labels.keys(), key=len, reverse=True)

    for name in sorted_names:
        label = names_to_labels[name]
        escaped = re.escape(name)
        # Replace possessives first (longer match), then plain names.
        # Lookbehind/lookahead for < > prevents matching inside existing labels.
        text = re.sub(
            r"(?<!<)\b" + escaped + r"'s\b",
            f"<{label}>'s",
            text,
        )
        text = re.sub(
            r"(?<!<)\b" + escaped + r"\b(?!>)",
            f"<{label}>",
            text,
        )

    return text


def anonymize_text(text: str) -> str:
    """Anonymize text by processing each line independently.

    Line-by-line processing prevents Presidio from detecting spans that cross
    line boundaries, which can eat markdown formatting between lines.
    After the NER pass, a cleanup pass catches any remaining name occurrences.
    """
    if not text or not text.strip():
        return text

    lines = text.split("\n")
    anonymized_lines = [anonymize_line(line) for line in lines]
    result = "\n".join(anonymized_lines)
    result = cleanup_pass(result)
    return result


def process_text_file(input_path: Path, output_path: Path):
    """Anonymize a plain text or markdown file."""
    text = input_path.read_text(encoding="utf-8")
    anonymized = anonymize_text(text)
    output_path.write_text(anonymized, encoding="utf-8")


def process_csv_file(input_path: Path, output_path: Path):
    """Anonymize every cell in a CSV file."""
    with open(input_path, "r", encoding="utf-8", newline="") as infile:
        reader = csv.reader(infile)
        rows = list(reader)

    anonymized_rows = []
    for row in rows:
        anonymized_rows.append([anonymize_text(cell) for cell in row])

    with open(output_path, "w", encoding="utf-8", newline="") as outfile:
        writer = csv.writer(outfile)
        writer.writerows(anonymized_rows)


def process_xlsx_file(input_path: Path, output_path: Path):
    """Anonymize every cell in an Excel file (all sheets)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ERROR: openpyxl not installed. Run: pip3 install openpyxl")
        return

    # Copy the file first to preserve formatting, then modify in place
    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = anonymize_text(cell.value)

    wb.save(output_path)


def anonymize_json_value(data):
    """Recursively anonymize all string values in a JSON structure."""
    if isinstance(data, str):
        return anonymize_text(data)
    elif isinstance(data, list):
        return [anonymize_json_value(item) for item in data]
    elif isinstance(data, dict):
        return {key: anonymize_json_value(value) for key, value in data.items()}
    else:
        return data


def process_json_file(input_path: Path, output_path: Path):
    """Anonymize all string values in a JSON file."""
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    anonymized = anonymize_json_value(data)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(anonymized, f, indent=2, ensure_ascii=False)


def main():
    # Create folders if they don't exist
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    # If filenames are passed as arguments, process only those.
    # Otherwise, process all supported files in input/.
    if len(sys.argv) > 1:
        files = []
        for name in sys.argv[1:]:
            path = INPUT_DIR / name
            if not path.exists():
                print(f"  File not found: {path}")
            elif path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                print(f"  Unsupported file type: {name}")
            else:
                files.append(path)
    else:
        files = [
            f for f in INPUT_DIR.iterdir()
            if f.is_file()
            and f.suffix.lower() in SUPPORTED_EXTENSIONS
            and not f.name.startswith(".")
        ]

    if not files:
        print(f"No files found in {INPUT_DIR}/")
        print(f"Drop .md, .txt, .csv, .xlsx, or .json files there and run again.")
        return

    print(f"Found {len(files)} file(s) to process:\n")

    for filepath in sorted(files):
        ext = filepath.suffix.lower()
        output_path = OUTPUT_DIR / filepath.name
        print(f"  Processing: {filepath.name} ...", end=" ", flush=True)

        try:
            if ext in (".md", ".txt"):
                process_text_file(filepath, output_path)
            elif ext == ".csv":
                process_csv_file(filepath, output_path)
            elif ext == ".xlsx":
                process_xlsx_file(filepath, output_path)
            elif ext == ".json":
                process_json_file(filepath, output_path)

            print("done")
        except Exception as e:
            print(f"ERROR: {e}")

    # Save the person mapping so you can see who is who
    if entity_map:
        map_path = OUTPUT_DIR / "entity_map.txt"
        with open(map_path, "w", encoding="utf-8") as f:
            f.write("Person Mapping\n")
            f.write("=" * 50 + "\n")
            # Collect unique labels and all names that map to each
            label_to_names = {}
            for (_, original), label in entity_map.items():
                label_to_names.setdefault(label, set()).add(original)
            for label in sorted(label_to_names, key=lambda l: int(l.split("_")[1])):
                names = sorted(label_to_names[label], key=len, reverse=True)
                f.write(f"  <{label}> = {', '.join(names)}\n")
        print(f"  Person mapping saved to: entity_map.txt")

    print(f"\nAnonymized files saved to: {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
