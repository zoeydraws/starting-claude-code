from __future__ import annotations

"""UOB (Singapore) Excel/CSV parser.

UOB format (based on actual exports):
- Files are typically Excel (.xls/.xlsx), not CSV
- Row 6, column 2 contains card/account type (used for detection)
- Card types: "UOB ONE CARD", "PREFERRED PLATINUM VISA", "LADY'S SOLITAIRE CARD", "One Account"
- For "One Account" (bank): data starts at row 9
- For credit cards: data starts at row 11
- Columns typically: Date, Description, Withdrawal/Debit, Deposit/Credit, Balance
"""

import csv
from pathlib import Path

from dateutil.parser import parse as parse_date

from ..models import Transaction
from .base import BaseParser


# Mapping from UOB account type to config bank identifier
UOB_ACCOUNT_TYPE_MAP = {
    "One Account": "uob-bank",
    "UOB ONE CARD": "uob-one-cc",
    "PREFERRED PLATINUM VISA": "uob-ppv-cc",
    "LADY'S SOLITAIRE CARD": "uob-ladies-cc",
}


class UOBParser(BaseParser):
    """Parser for UOB Singapore Excel/CSV exports."""

    bank_name = "UOB"
    detection_columns = []

    # Known UOB card/account types
    ACCOUNT_TYPES = list(UOB_ACCOUNT_TYPE_MAP.keys())

    def detect(self, file_path: str | Path) -> bool:
        """Detect UOB file by checking for account type in row 6, col 2."""
        file_path = Path(file_path)

        # Check file extension
        if file_path.suffix.lower() in [".xls", ".xlsx"]:
            return self._detect_excel(file_path)
        elif file_path.suffix.lower() == ".csv":
            return self._detect_csv(file_path)
        return False

    def get_account_type(self, file_path: str | Path) -> str | None:
        """Get the specific UOB account type from the file.

        Returns the normalized account type identifier (e.g., 'uob-bank', 'uob-one-cc')
        or None if not detected.
        """
        file_path = Path(file_path)
        raw_type = self._get_raw_account_type(file_path)

        if raw_type and raw_type in UOB_ACCOUNT_TYPE_MAP:
            return UOB_ACCOUNT_TYPE_MAP[raw_type]
        return None

    def get_raw_account_type(self, file_path: str | Path) -> str | None:
        """Get the raw account type string from the file (e.g., 'UOB ONE CARD')."""
        return self._get_raw_account_type(Path(file_path))

    def _get_raw_account_type(self, file_path: Path) -> str | None:
        """Internal method to get raw account type."""
        if file_path.suffix.lower() in [".xls", ".xlsx"]:
            return self._get_account_type_excel(file_path)
        elif file_path.suffix.lower() == ".csv":
            return self._get_account_type_csv(file_path)
        return None

    def _get_account_type_excel(self, file_path: Path) -> str | None:
        """Get account type from Excel file."""
        if file_path.suffix.lower() == ".xls":
            return self._get_account_type_xls(file_path)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            cell_value = sheet.cell(row=6, column=2).value
            wb.close()
            if cell_value:
                return str(cell_value).strip()
        except Exception:
            pass
        return None

    def _get_account_type_xls(self, file_path: Path) -> str | None:
        """Get account type from old-format .xls file using xlrd."""
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            if sheet.nrows > 5 and sheet.ncols > 1:
                cell_value = sheet.cell_value(5, 1)
                if cell_value:
                    return str(cell_value).strip()
        except Exception:
            pass
        return None

    def _get_account_type_csv(self, file_path: Path) -> str | None:
        """Get account type from CSV file."""
        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
                if len(rows) > 5 and len(rows[5]) > 1:
                    return rows[5][1].strip()
        except Exception:
            pass
        return None

    def _detect_excel(self, file_path: Path) -> bool:
        """Detect UOB Excel file."""
        account_type = self._get_account_type_excel(file_path)
        return account_type in self.ACCOUNT_TYPES if account_type else False

    def _detect_csv(self, file_path: Path) -> bool:
        """Detect UOB CSV file (fallback if converted from Excel)."""
        account_type = self._get_account_type_csv(file_path)
        return account_type in self.ACCOUNT_TYPES if account_type else False

    def parse(self, file_path: str | Path) -> list[Transaction]:
        """Parse UOB file (Excel or CSV)."""
        file_path = Path(file_path)

        if file_path.suffix.lower() in [".xls", ".xlsx"]:
            return self._parse_excel(file_path)
        else:
            return self._parse_csv(file_path)

    def _parse_excel(self, file_path: Path) -> list[Transaction]:
        """Parse UOB Excel file (.xls or .xlsx)."""
        if file_path.suffix.lower() == ".xls":
            return self._parse_xls(file_path)

        transactions = []

        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required to parse UOB .xlsx files. "
                "Install with: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active

        # Get all rows as list
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 9:
            return transactions

        # Determine account type and data start row
        account_type = rows[5][1] if len(rows) > 5 and len(rows[5]) > 1 else ""

        if account_type == "One Account":
            data_start = 8  # Row 9 (0-indexed: 8)
            header_row = 7  # Row 8
        else:
            data_start = 10  # Row 11 (0-indexed: 10)
            header_row = 9  # Row 10

        if len(rows) <= data_start:
            return transactions

        # Parse header
        header = rows[header_row] if len(rows) > header_row else []
        header_lower = [str(h).lower().strip() if h else "" for h in header]

        # Find column indices
        date_idx = self._find_column(header_lower, ["transaction date", "date"])
        desc_idx = self._find_column(header_lower, ["description", "transaction description", "details"])
        debit_idx = self._find_column(header_lower, ["withdrawal", "debit"])
        credit_idx = self._find_column(header_lower, ["deposit", "credit"])
        # CC statements use a single "Transaction Amount(Local)" column
        amount_idx = self._find_column(header_lower, ["transaction amount(local)"])
        is_cc = account_type != "One Account"

        # Process transaction rows
        for row in rows[data_start:]:
            if not row:
                continue

            # Parse date
            date_val = row[date_idx] if date_idx is not None and date_idx < len(row) else None
            if not date_val:
                continue

            try:
                if hasattr(date_val, "date"):
                    # Already a datetime object
                    txn_date = date_val.date() if hasattr(date_val, "date") else date_val
                else:
                    # String date
                    date_str = str(date_val).strip().replace('"', '')
                    txn_date = parse_date(date_str, dayfirst=True).date()
            except (ValueError, TypeError, AttributeError):
                continue

            # Parse amount
            amount = self._parse_amount_from_row(
                row, debit_idx, credit_idx, amount_idx, is_cc
            )

            if amount == 0:
                continue

            # Get description
            payee = ""
            if desc_idx is not None and desc_idx < len(row) and row[desc_idx]:
                payee = str(row[desc_idx]).strip()
            if not payee:
                payee = "UOB Transaction"
            payee = payee[:100]

            transactions.append(Transaction.from_amount(
                date=txn_date,
                amount=amount,
                payee_name=payee,
            ))

        return transactions

    def _parse_xls(self, file_path: Path) -> list[Transaction]:
        """Parse old-format .xls file using xlrd."""
        transactions = []

        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd is required to parse UOB .xls files. "
                "Install with: pip install xlrd"
            )

        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)

        # Convert to list of tuples for consistency with openpyxl path
        rows = []
        for i in range(sheet.nrows):
            rows.append(tuple(sheet.cell_value(i, j) for j in range(sheet.ncols)))

        if len(rows) < 9:
            return transactions

        # Determine account type and data start row
        account_type = rows[5][1] if len(rows) > 5 and len(rows[5]) > 1 else ""

        if account_type == "One Account":
            data_start = 8  # Row 9 (0-indexed: 8)
            header_row = 7  # Row 8
        else:
            data_start = 10  # Row 11 (0-indexed: 10)
            header_row = 9  # Row 10

        if len(rows) <= data_start:
            return transactions

        # Parse header
        header = rows[header_row] if len(rows) > header_row else ()
        header_lower = [str(h).lower().strip() if h else "" for h in header]

        # Find column indices
        date_idx = self._find_column(header_lower, ["transaction date", "date"])
        desc_idx = self._find_column(header_lower, ["description", "transaction description", "details"])
        debit_idx = self._find_column(header_lower, ["withdrawal", "debit"])
        credit_idx = self._find_column(header_lower, ["deposit", "credit"])
        # CC statements use a single "Transaction Amount(Local)" column
        amount_idx = self._find_column(header_lower, ["transaction amount(local)"])
        is_cc = account_type != "One Account"

        # Process transaction rows
        for row in rows[data_start:]:
            if not row:
                continue

            # Parse date
            date_val = row[date_idx] if date_idx is not None and date_idx < len(row) else None
            if not date_val:
                continue

            try:
                if isinstance(date_val, float):
                    # xlrd may return dates as floats - convert via xlrd
                    date_tuple = xlrd.xldate_as_tuple(date_val, wb.datemode)
                    from datetime import date
                    txn_date = date(date_tuple[0], date_tuple[1], date_tuple[2])
                else:
                    date_str = str(date_val).strip()
                    txn_date = parse_date(date_str, dayfirst=True).date()
            except (ValueError, TypeError, AttributeError):
                continue

            # Parse amount
            amount = self._parse_amount_from_row(
                row, debit_idx, credit_idx, amount_idx, is_cc
            )

            if amount == 0:
                continue

            # Get description
            payee = ""
            if desc_idx is not None and desc_idx < len(row) and row[desc_idx]:
                payee = str(row[desc_idx]).strip()
            if not payee:
                payee = "UOB Transaction"
            payee = payee[:100]

            transactions.append(Transaction.from_amount(
                date=txn_date,
                amount=amount,
                payee_name=payee,
            ))

        return transactions

    def _parse_csv(self, file_path: Path) -> list[Transaction]:
        """Parse UOB CSV file (converted from Excel)."""
        transactions = []

        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if len(rows) < 9:
            return transactions

        # Determine account type and data start row
        account_type = rows[5][1] if len(rows) > 5 and len(rows[5]) > 1 else ""

        if account_type == "One Account":
            data_start = 8
            header_row = 7
        else:
            data_start = 10
            header_row = 9

        if len(rows) <= data_start:
            return transactions

        # Parse header
        header = rows[header_row] if len(rows) > header_row else []
        header_lower = [h.lower().strip() for h in header]

        # Find column indices
        date_idx = self._find_column(header_lower, ["transaction date", "date"])
        desc_idx = self._find_column(header_lower, ["description", "transaction description", "details"])
        debit_idx = self._find_column(header_lower, ["withdrawal", "debit"])
        credit_idx = self._find_column(header_lower, ["deposit", "credit"])

        # Process transaction rows
        for row in rows[data_start:]:
            if not row:
                continue

            # Parse date
            date_str = row[date_idx].strip().replace('"', '') if date_idx is not None and date_idx < len(row) else ""
            if not date_str:
                continue

            try:
                txn_date = parse_date(date_str, dayfirst=True).date()
            except (ValueError, TypeError):
                continue

            # Parse amount
            amount = 0.0

            if debit_idx is not None and debit_idx < len(row):
                debit_str = row[debit_idx].replace(",", "").replace("$", "").replace('"', '').strip()
                if debit_str:
                    try:
                        amount = -abs(float(debit_str))
                    except ValueError:
                        pass

            if amount == 0 and credit_idx is not None and credit_idx < len(row):
                credit_str = row[credit_idx].replace(",", "").replace("$", "").replace('"', '').strip()
                if credit_str:
                    try:
                        amount = abs(float(credit_str))
                    except ValueError:
                        pass

            if amount == 0:
                continue

            # Get description
            payee = ""
            if desc_idx is not None and desc_idx < len(row):
                payee = row[desc_idx].strip().replace('"', '')
            if not payee:
                payee = "UOB Transaction"
            payee = payee[:100]

            transactions.append(Transaction.from_amount(
                date=txn_date,
                amount=amount,
                payee_name=payee,
            ))

        return transactions

    def _parse_amount_from_row(
        self, row, debit_idx, credit_idx, amount_idx, is_cc: bool
    ) -> float:
        """Parse amount from a transaction row.

        Handles both separate debit/credit columns (bank accounts)
        and single transaction amount column (credit cards).
        """
        amount = 0.0

        # Try debit/credit columns first
        if debit_idx is not None and debit_idx < len(row) and row[debit_idx]:
            try:
                debit_val = row[debit_idx]
                if isinstance(debit_val, (int, float)) and debit_val != 0:
                    amount = -abs(float(debit_val))
                elif isinstance(debit_val, str):
                    debit_str = debit_val.replace(",", "").replace("$", "").strip()
                    if debit_str:
                        amount = -abs(float(debit_str))
            except (ValueError, TypeError):
                pass

        if amount == 0 and credit_idx is not None and credit_idx < len(row) and row[credit_idx]:
            try:
                credit_val = row[credit_idx]
                if isinstance(credit_val, (int, float)) and credit_val != 0:
                    amount = abs(float(credit_val))
                elif isinstance(credit_val, str):
                    credit_str = credit_val.replace(",", "").replace("$", "").strip()
                    if credit_str:
                        amount = abs(float(credit_str))
            except (ValueError, TypeError):
                pass

        # Fallback to single amount column (CC statements)
        if amount == 0 and amount_idx is not None and amount_idx < len(row) and row[amount_idx]:
            try:
                amt_val = row[amount_idx]
                if isinstance(amt_val, (int, float)) and amt_val != 0:
                    # For CC, positive = expense (outflow)
                    amount = -abs(float(amt_val)) if is_cc else float(amt_val)
                elif isinstance(amt_val, str):
                    amt_str = amt_val.replace(",", "").replace("$", "").strip()
                    if amt_str:
                        amount = -abs(float(amt_str)) if is_cc else float(amt_str)
            except (ValueError, TypeError):
                pass

        return amount

    def _find_column(self, headers: list[str], names: list[str]) -> int | None:
        """Find column index matching any of the given names."""
        for i, header in enumerate(headers):
            for name in names:
                if name in header:
                    return i
        return None
